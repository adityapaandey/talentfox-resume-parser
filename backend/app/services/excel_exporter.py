import pandas as pd
from typing import List, Dict
from datetime import datetime
import os

class ExcelExporter:
    """Export parsed resume data to Excel."""
    
    def export_to_excel(self, parsed_resumes: List[Dict], output_path: str) -> str:
        """Export list of parsed resumes to Excel file."""
        try:
            # Convert to DataFrame
            df = self._create_dataframe(parsed_resumes)
            
            # Create Excel file with formatting
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Resumes', index=False)
                
                # Get workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Resumes']
                
                # Format worksheet
                self._format_worksheet(worksheet, df)
            
            return output_path
            
        except Exception as e:
            raise Exception(f"Error exporting to Excel: {str(e)}")
    
    def _create_dataframe(self, parsed_resumes: List[Dict]) -> pd.DataFrame:
        """Create DataFrame from parsed resume data."""
        # Define columns in desired order
        columns = [
            'name', 'email', 'phone', 'experience_years',
            'skills', 'education', 'linkedin', 'github',
            'summary', 'filename', 'parsed_date'
        ]
        
        # Prepare data
        data = []
        for resume in parsed_resumes:
            row = {}
            for col in columns:
                value = resume.get(col)
                
                # Convert lists to comma-separated strings
                if isinstance(value, list):
                    row[col] = ', '.join(value) if value else ''
                else:
                    row[col] = value if value is not None else ''
            
            data.append(row)
        
        # Create DataFrame
        df = pd.DataFrame(data, columns=columns)
        
        # Rename columns for better readability
        df.columns = [
            'Name', 'Email', 'Phone', 'Experience',
            'Skills', 'Education', 'LinkedIn', 'GitHub',
            'Summary', 'Filename', 'Parsed Date'
        ]
        
        return df
    
    def _format_worksheet(self, worksheet, df):
        """Apply formatting to worksheet."""
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        
        # Format header row
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths
        column_widths = {
            'A': 25,  # Name
            'B': 30,  # Email
            'C': 18,  # Phone
            'D': 15,  # Experience
            'E': 50,  # Skills
            'F': 40,  # Education
            'G': 35,  # LinkedIn
            'H': 35,  # GitHub
            'I': 60,  # Summary
            'J': 30,  # Filename
            'K': 20,  # Parsed Date
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Wrap text for summary column
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=9, max_col=9):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Freeze header row
        worksheet.freeze_panes = 'A2'
